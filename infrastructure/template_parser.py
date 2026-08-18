"""
Парсер шаблона цен Ozon.
Извлекает данные из листа "Товары и цены", рассчитывает реальную цену.
"""

import zipfile
from pathlib import Path
from xml.etree import ElementTree as ET

import pandas as pd
from openpyxl import load_workbook

from infrastructure.logger import logger

# Константы для парсинга
MIN_SHEET_INDEX = 1  # Второй лист (индекс 1)
MIN_ROWS_REQUIRED = 2  # Минимум строк в листе (заголовок + данные)


class TemplateParser:
    """
    Класс для парсинга и расчёта реальной цены из шаблона Ozon.
    """

    COLUMN_NAMES = {
        "sku": "SKU",
        "status": "Статус",
        "visibility": "Видимость на OZON",
        "stock_ozon": "На складе Ozon",
        "stock_my": "На моих складах",
        "price_before_discount": "Цена до скидки, руб.",
        "price_with_discount": "Цена с учетом акции или стратегии, руб.",
        "discount_with_action": "Скидка с учетом акции, руб.",
        "acquiring": "Эквайринг",
        "fbo_reward_percent": "Вознаграждение Ozon, FBO, %",
        "fbo_logistics_min": "Логистика Ozon, минимум, FBO",
        "fbo_logistics_max": "Логистика Ozon, максимум, FBO",
        "fbo_delivery": "Доставка до места выдачи, FBO",
        "fbo_handling": "Обработка нестандартного товара, FBO",
        "fbs_reward_percent": "Вознаграждение Ozon, FBS, %",
        "fbs_handling_min": "Обработка отправления, минимум FBS",
        "fbs_handling_max": "Обработка отправления, максимум FBS",
        "fbs_handling_nonstandard": "Обработка нестандартного товара, FBS",
        "fbs_logistics_min": "Логистика Ozon, минимум, FBS",
        "fbs_logistics_max": "Логистика Ozon, максимум, FBS",
        "fbs_delivery": "Доставка до места выдачи, FBS",
    }

    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.df: pd.DataFrame | None = None

    def _read_xlsx_via_zip(self) -> pd.DataFrame | None:
        """
        Читает xlsx как zip-архив, извлекая данные напрямую из XML,
        минуя повреждённые стили.
        """
        try:
            with zipfile.ZipFile(self.file_path, "r") as zf:
                sheet_files = self._get_sheet_files(zf)
                target_sheet = self._find_target_sheet(zf, sheet_files)
                if not target_sheet:
                    return None

                shared_strings = self._read_shared_strings(zf)
                return self._parse_sheet(zf, target_sheet, shared_strings)

        except Exception as e:
            logger.error(f"Ошибка при чтении через zip: {e}")
            return None

    def _get_sheet_files(self, zf: zipfile.ZipFile) -> list[str]:
        """Получает отсортированный список файлов листов."""
        sheet_files = [
            f for f in zf.namelist() if f.startswith("xl/worksheets/sheet") and f.endswith(".xml")
        ]
        sheet_files.sort()
        logger.info(f"Найдены листы: {sheet_files}")
        return sheet_files

    def _find_target_sheet(self, zf: zipfile.ZipFile, sheet_files: list[str]) -> str | None:
        """Находит целевой лист с данными."""
        for sf in sheet_files:
            try:
                content = zf.read(sf).decode("utf-8", errors="ignore")
                if "SKU" in content and "Статус" in content:
                    logger.info(f"Найден лист с данными: {sf}")
                    return sf
            except Exception:
                continue

        if len(sheet_files) > MIN_SHEET_INDEX:
            target_sheet = sheet_files[MIN_SHEET_INDEX]
            logger.info(f"Используем второй лист: {target_sheet}")
            return target_sheet

        logger.error("Не найден лист с данными")
        return None

    def _read_shared_strings(self, zf: zipfile.ZipFile) -> list[str]:
        """Читает shared strings из xlsx."""
        shared_strings = []
        try:
            ss_xml = zf.read("xl/sharedStrings.xml")
            ss_root = ET.fromstring(ss_xml)
            for si in ss_root.findall(
                ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}si"
            ):
                texts = [
                    t.text or ""
                    for t in si.findall(
                        ".//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}t"
                    )
                ]
                shared_strings.append("".join(texts))
            logger.info(f"Загружено {len(shared_strings)} общих строк")
        except Exception as e:
            logger.warning(f"Не удалось прочитать sharedStrings: {e}")
        return shared_strings

    def _parse_sheet(
        self, zf: zipfile.ZipFile, target_sheet: str, shared_strings: list[str]
    ) -> pd.DataFrame | None:
        """Парсит целевой лист в DataFrame."""
        sheet_xml = zf.read(target_sheet)
        root = ET.fromstring(sheet_xml)
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        rows = []
        for row_elem in root.findall(".//main:row", ns):
            row_data = []
            for cell in row_elem.findall(".//main:c", ns):
                cell_type = cell.get("t")
                value_elem = cell.find(".//main:v", ns)
                if value_elem is not None:
                    value = value_elem.text
                    if cell_type == "s" and value is not None:
                        idx = int(value)
                        value = shared_strings[idx] if idx < len(shared_strings) else ""
                else:
                    value = ""
                row_data.append(value)
            if any(row_data):  # пропускаем пустые строки
                rows.append(row_data)

        if len(rows) < MIN_ROWS_REQUIRED:
            logger.error("Недостаточно данных в листе")
            return None

        headers = rows[1]  # вторая строка
        # Удаляем пустые заголовки справа
        while headers and not headers[-1]:
            headers.pop()

        data = rows[2:]  # данные начиная с третьей строки
        data = [row[: len(headers)] for row in data]  # обрезаем до длины заголовков

        df = pd.DataFrame(data, columns=headers)
        logger.info(f"Из zip-архива получено {len(df)} строк")
        return df

    def load(self) -> bool:
        """
        Загружает Excel-файл, сначала через openpyxl, при ошибке через zip.
        """
        if self._load_via_openpyxl():
            return self._post_load_processing()

        logger.info("Пробуем прямой разбор zip-архива...")
        if self._load_via_zip():
            return self._post_load_processing()

        logger.error("Не удалось прочитать файл")
        return False

    def _load_via_openpyxl(self) -> bool:
        """Пытается загрузить файл через openpyxl."""
        try:
            wb = load_workbook(self.file_path, data_only=True, read_only=True, keep_links=False)
            if "Товары и цены" not in wb.sheetnames:
                logger.error("Лист 'Товары и цены' не найден")
                wb.close()
                return False
            ws = wb["Товары и цены"]
            data = []
            headers = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i == 0:
                    continue  # первая строка служебная
                if i == 1:
                    headers = [str(cell).strip() if cell is not None else "" for cell in row]
                    while headers and not headers[-1]:
                        headers.pop()
                    continue
                if any(cell is not None for cell in row):
                    data.append(row[: len(headers)])
            wb.close()
            self.df = pd.DataFrame(data, columns=headers)
            logger.info(f"Загружен через openpyxl, строк: {len(self.df)}")
            return True
        except Exception as e:
            logger.error(f"Ошибка openpyxl: {e}")
            return False

    def _load_via_zip(self) -> bool:
        """Пытается загрузить файл через zip-парсинг."""
        self.df = self._read_xlsx_via_zip()
        if self.df is None:
            logger.error("Не удалось прочитать файл")
            return False
        logger.info(f"Загружен через zip, строк: {len(self.df)}")
        return True

    def _post_load_processing(self) -> bool:
        """Выполняет пост-обработку после загрузки: валидация и конвертация."""
        if not self._validate_columns():
            return False
        self._convert_numeric_columns()
        logger.info("��� Файл успешно загружен и подготовлен")
        return True

    def _validate_columns(self) -> bool:
        """Проверяет наличие всех необходимых столбцов."""
        if self.df is None:
            logger.error("DataFrame не загружен")
            return False
        missing = []
        for _key, col_name in self.COLUMN_NAMES.items():
            if col_name not in self.df.columns:
                missing.append(col_name)
        if missing:
            logger.error(f"Не найдены столбцы: {missing}")
            return False
        return True

    def _convert_numeric_columns(self) -> None:
        """Преобразует числовые столбцы."""
        if self.df is None:
            logger.error("DataFrame не загружен")
            return

        numeric_cols = [
            "stock_ozon",
            "stock_my",
            "price_before_discount",
            "price_with_discount",
            "discount_with_action",
            "acquiring",
            "fbo_reward_percent",
            "fbo_logistics_min",
            "fbo_logistics_max",
            "fbo_delivery",
            "fbo_handling",
            "fbs_reward_percent",
            "fbs_handling_min",
            "fbs_handling_max",
            "fbs_handling_nonstandard",
            "fbs_logistics_min",
            "fbs_logistics_max",
            "fbs_delivery",
        ]
        for col in numeric_cols:
            if col in self.COLUMN_NAMES:
                col_name = self.COLUMN_NAMES[col]
                self.df[col_name] = pd.to_numeric(self.df[col_name], errors="coerce")

    def get_relevant_products(self) -> list[dict]:
        if self.df is None:
            logger.error("Данные не загружены")
            return []

        status_col = self.COLUMN_NAMES["status"]
        visibility_col = self.COLUMN_NAMES["visibility"]

        mask = (self.df[status_col] == "Продается") & (self.df[visibility_col].str.lower() == "да")
        filtered = self.df[mask].copy()

        if filtered.empty:
            logger.warning("Нет товаров, соответствующих критериям")
            return []

        products = []
        for _, row in filtered.iterrows():
            product = {}
            for key, col_name in self.COLUMN_NAMES.items():
                product[key] = row[col_name]
            products.append(product)

        logger.info(f"Найдено {len(products)} товаров для расчёта")
        return products

    def determine_sales_type(self, product: dict) -> str:
        stock_ozon = product.get("stock_ozon", 0)
        stock_my = product.get("stock_my", 0)

        if pd.isna(stock_ozon):
            stock_ozon = 0
        if pd.isna(stock_my):
            stock_my = 0

        if stock_ozon > 0 and stock_my > 0:
            return "mixed"
        if stock_ozon > 0:
            return "fbo"
        if stock_my > 0:
            return "fbs"
        return "none"

    def calculate_real_price(self, product: dict) -> float | None:
        sales_type = self.determine_sales_type(product)
        if sales_type == "none":
            return None

        price_before = product.get("price_before_discount", 0)
        discount_with_action = product.get("discount_with_action", 0)
        base_price = price_before - discount_with_action

        if sales_type in ("fbo", "fbs"):
            return base_price

        if sales_type == "mixed":
            price_with_discount = product.get("price_with_discount", 0)
            acquiring = product.get("acquiring", 0)
            fbo_reward = product.get("fbo_reward_percent", 0) / 100
            fbo_log_min = product.get("fbo_logistics_min", 0)
            fbo_log_max = product.get("fbo_logistics_max", 0)
            fbo_delivery = product.get("fbo_delivery", 0)
            fbo_handling = product.get("fbo_handling", 0)

            fbs_reward = product.get("fbs_reward_percent", 0) / 100
            fbs_hand_min = product.get("fbs_handling_min", 0)
            fbs_hand_max = product.get("fbs_handling_max", 0)
            fbs_hand_nonstd = product.get("fbs_handling_nonstandard", 0)
            fbs_log_min = product.get("fbs_logistics_min", 0)
            fbs_log_max = product.get("fbs_logistics_max", 0)
            fbs_delivery = product.get("fbs_delivery", 0)

            numerator = (
                acquiring
                + (fbo_reward * price_with_discount)
                + ((fbo_log_min + fbo_log_max) / 2)
                + fbo_delivery
                + fbo_handling
            )
            denominator = (
                (fbs_reward * price_with_discount)
                + ((fbs_hand_min + fbs_hand_max) / 2)
                + fbs_hand_nonstd
                + ((fbs_log_min + fbs_log_max) / 2)
                + fbs_delivery
            )

            if denominator == 0:
                logger.warning(f"Нулевой знаменатель для SKU {product.get('sku')}")
                return None

            real_price = base_price * (numerator / denominator)
            return round(real_price)
        return None

    def process(self) -> list[tuple[dict, float | None, str]]:
        products = self.get_relevant_products()
        results = []
        for p in products:
            sales_type = self.determine_sales_type(p)
            if sales_type == "none":
                continue
            real_price = self.calculate_real_price(p)
            results.append((p, real_price, sales_type))
        return results

    # ------------------------------------------------------------------
    # Вспомогательные методы для _read_xlsx_via_zip
    # ------------------------------------------------------------------

    def _read_sheet_data(
        self, zf: zipfile.ZipFile, target_sheet: str, shared_strings: list[str]
    ) -> list[list[str]]:
        """Читает данные листа из zip-архива."""
        sheet_xml = zf.read(target_sheet)
        root = ET.fromstring(sheet_xml)
        ns = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

        rows = []
        for row_elem in root.findall(".//main:row", ns):
            row_data = []
            for cell in row_elem.findall(".//main:c", ns):
                cell_type = cell.get("t")
                value_elem = cell.find(".//main:v", ns)
                if value_elem is not None:
                    value = value_elem.text
                    if cell_type == "s" and value is not None:
                        idx = int(value)
                        value = shared_strings[idx] if idx < len(shared_strings) else ""
                else:
                    value = ""
                row_data.append(value)
            if any(row_data):  # пропускаем пустые строки
                rows.append(row_data)
        return rows

    def _build_dataframe(self, rows: list[list[str]]) -> pd.DataFrame:
        """Создает DataFrame из строк данных."""
        headers = rows[1]  # вторая строка
        # Удаляем пустые заголовки справа
        while headers and not headers[-1]:
            headers.pop()

        data = rows[2:]  # данные начиная с третьей строки
        data = [row[: len(headers)] for row in data]  # обрезаем до длины заголовков

        return pd.DataFrame(data, columns=headers)

