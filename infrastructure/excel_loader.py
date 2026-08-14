"""
Загрузчик данных из Excel-файла.

Реализует интерфейс ILoader: читает товары, стратегии, цены конкурентов,
обновляет ячейки в файле после расчёта.
"""

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook

from config.settings import PERCENT_MAX, TIME_FORMAT_LENGTH, settings
from core.entities import PriceCalculationResult, ProductInfo, StrategyInterval
from core.enums import StrategyType, parse_strategy_value
from core.repository import ILoader
from infrastructure.logger import logger


class ExcelLoader(ILoader):
    """
    Загрузчик данных из Excel-файла.

    Читает:
        - товары (SKU, себестоимость, РИЦ, старая цена),
        - цены конкурентов (до 5),
        - интервалы стратегий (до N, настраивается через SCHEDULE_INTERVALS_COUNT).

    После расчёта обновляет ячейки с ценами и маржинальностью,
    сохраняя форматирование через openpyxl.
    """

    # Маппинг названий колонок для обратной совместимости
    COLUMN_MAPPING = {
        "sku": ["sku", "артикул", "article"],
        "cost_price": ["себестоимость", "cost_price", "cost"],
        "min_price": ["цена риц"],
        "old_price": ["цена до скидки", "old_price", "старая цена"],
    }

    def __init__(self, file_path: Path) -> None:
        """
        Инициализирует загрузчик.

        Args:
            file_path: Путь к Excel-файлу.
        """
        self.file_path = file_path
        self._strategies: dict[str, list[StrategyInterval]] = {}

    # ------------------------------------------------------------------
    # Реализация интерфейса ILoader
    # ------------------------------------------------------------------

    def load(self) -> tuple[list[ProductInfo], list[str]]:
        """
        Загружает товары из Excel с валидацией.

        Returns:
            Кортеж (список товаров, список предупреждений/ошибок).

        Raises:
            Возвращает пустой список и сообщение об ошибке, если файл не найден,
            имеет неверный формат, отсутствует колонка SKU или есть дубликаты SKU.
        """
        # Валидация файла
        error = self._validate_file()
        if error:
            return [], [error]

        df = pd.read_excel(self.file_path, engine="openpyxl", dtype=str)
        df.columns = df.columns.str.lower().str.strip()

        # Поиск колонки SKU
        sku_col = self._find_sku_column(df)
        if sku_col is None:
            return [], ["Не найдена колонка SKU (ожидаются: sku, артикул, article, id, offer_id)"]

        # Проверка дубликатов SKU
        error = self._check_duplicates(df, sku_col)
        if error:
            return [], [error]

        products = []
        warnings = []
        self._strategies.clear()

        for i, (_, row) in enumerate(df.iterrows(), start=2):
            product, row_warnings = self._parse_row(row, df.columns, sku_col, i)
            if product:
                products.append(product)
            warnings.extend(row_warnings)

        logger.info(f"Загружено {len(products)} товаров, {len(warnings)} предупреждений")
        return products, warnings

    def _validate_file(self) -> str | None:
        """Валидирует существование и формат файла."""
        if not self.file_path.exists():
            logger.error(f"Файл {self.file_path} не найден")
            return "Файл Excel не найден"

        if self.file_path.suffix.lower() != ".xlsx":
            logger.error(f"Неподдерживаемый формат: {self.file_path.suffix}. Используйте .xlsx")
            return f"Неподдерживаемый формат: {self.file_path.suffix}"
        return None

    def _find_sku_column(self, df: pd.DataFrame) -> str | None:
        """Находит колонку SKU в DataFrame."""
        for col in ["sku", "артикул", "article", "id", "offer_id"]:
            if col in df.columns:
                return col
        return None

    def _check_duplicates(self, df: pd.DataFrame, sku_col: str) -> str | None:
        """Проверяет дубликаты SKU."""
        df["_sku_normalized"] = df[sku_col].astype(str).str.strip()
        duplicates = df[df["_sku_normalized"].duplicated(keep=False)]["_sku_normalized"].unique()
        if len(duplicates) > 0:
            return f"Обнаружены дубликаты SKU: {', '.join(duplicates)}"
        return None

    def _parse_row(
        self, row: pd.Series, columns: pd.Index, sku_col: str, row_num: int
    ) -> tuple[ProductInfo | None, list[str]]:
        """Парсит строку DataFrame в ProductInfo."""
        warnings = []
        sku = str(row[sku_col]).strip()
        if not sku:
            warnings.append(f"Строка {row_num}: пропущен SKU")
            return None, warnings

        # Валидация себестоимости
        cost_price = self._get_float(row, columns, ["себестоимость", "cost_price", "cost"], 0.0)
        if cost_price <= 0:
            warnings.append(f"SKU {sku}: себестоимость = {cost_price} <= 0, товар пропущен")
            return None, warnings

        min_price = self._get_float(row, columns, ["цена риц", "min_price", "rip"], 0.0)

        # Чтение цен конкурентов
        competitor_min_price = self._read_competitor_prices(row, columns)

        # Парсинг интервалов стратегий
        intervals, interval_warnings = self._parse_intervals_with_validation(row, columns)
        warnings.extend(interval_warnings)

        if not intervals:
            warnings.append(
                f"SKU {sku}: не задано ни одного интервала стратегии, "
                "используется стратегия по умолчанию 'Равная'"
            )
            intervals = self._get_default_intervals()

        old_price_val = self._get_float(
            row, columns, ["цена до скидки", "old_price", "старая цена"], 0.0
        )
        old_price = old_price_val if old_price_val > 0 else None

        product = ProductInfo(
            sku=sku,
            product_name=None,
            cost_price=cost_price,
            min_price=min_price,
            current_price=0.0,
            old_price=old_price,
            competitor_min_price=competitor_min_price,
        )
        self._strategies[sku] = intervals
        return product, warnings

    def _read_competitor_prices(self, row: pd.Series, columns: pd.Index) -> float | None:
        """Читает цены конкурентов из строки."""
        competitor_prices = []
        price_prefix = settings.COMPETITOR_PRICE_COLUMN_PREFIX
        for j in range(1, settings.MAX_COMPETITORS + 1):
            price_col = f"{price_prefix} {j}"
            if price_col in columns:
                val = row.get(price_col)
                if pd.notna(val):
                    try:
                        price = float(val)
                        if price > 0:
                            competitor_prices.append(price)
                    except (ValueError, TypeError):
                        pass
        return min(competitor_prices) if competitor_prices else None

    def _get_default_intervals(self) -> list[StrategyInterval]:
        """Возвращает интервалы стратегии по умолчанию."""
        return [
            StrategyInterval(
                start="00:00", end="23:59", strategy_type=StrategyType.EQUAL, percent=0.0
            )
        ]

    def get_strategy_intervals(self, product: ProductInfo) -> list[StrategyInterval]:
        """
        Возвращает интервалы стратегий для товара (из загруженных данных).

        Args:
            product: Объект товара (используется SKU).

        Returns:
            Список StrategyInterval (пустой, если не найдено).
        """
        return self._strategies.get(product.sku, [])

    def update_product_in_file(self, sku: str, updates: dict[str, Any]) -> bool:
        """
        Обновляет данные товара в Excel-файле (точечное обновление ячеек).

        Args:
            sku: Артикул товара.
            updates: Словарь {поле: значение}.

        Returns:
            True в случае успеха, False при ошибке.
        """
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            if ws is None:
                logger.error("Не удалось получить активный лист")
                return False

            header_row = 1
            col_map = self._build_column_map(ws, header_row)

            # Поиск колонки SKU
            sku_col = self._find_sku_column_in_sheet(ws, header_row)
            if sku_col is None:
                logger.error("Не найдена колонка 'SKU'")
                return False

            # Поиск строки с нужным SKU
            target_row = self._find_target_row(ws, sku_col, sku)
            if target_row is None:
                logger.warning(f"SKU {sku} не найден в файле")
                return False

            # Обновляем ячейки
            self._update_cells(ws, target_row, col_map, updates)

            wb.save(self.file_path)
            logger.info(f"Обновлены поля {list(updates.keys())} для SKU {sku}")
            return True
        except Exception as e:
            logger.error(f"Ошибка обновления Excel: {e}")
            return False

    def _build_column_map(self, ws, header_row: int) -> dict[str, int]:
        """Строит маппинг полей на индексы колонок."""
        col_map = {}
        target_columns = {
            "current_price": ["ваша цена", "current_price", "price"],
            "min_price": ["минимальная цена", "min_price", "min"],
            "old_price": ["цена до скидки", "old_price", "старая цена"],
            "margin": ["маржинальность", "маржа", "margin"],
            "margin_week": ["маржинальность за неделю", "margin_week"],
            "margin_month": ["маржинальность за месяц", "margin_month"],
            "product_name": ["название", "name", "товар", "product_name"],
        }

        for field, synonyms in target_columns.items():
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value and str(cell.value).lower().strip() in synonyms:
                    col_map[field] = col_idx
                    break
        return col_map

    def _find_sku_column_in_sheet(self, ws, header_row: int) -> int | None:
        """Находит колонку SKU в листе."""
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value and str(cell.value).lower().strip() in ["sku", "артикул", "offer_id"]:
                return col_idx
        return None

    def _find_target_row(self, ws, sku_col: int, sku: str) -> int | None:
        """Находит строку с заданным SKU."""
        for row_idx in range(2, ws.max_row + 1):
            sku_cell = ws.cell(row_idx, sku_col)
            cell_value = sku_cell.value
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip()
            if cell_str == str(sku):
                return row_idx
            try:
                if int(float(cell_str)) == int(float(sku)):
                    return row_idx
            except (ValueError, TypeError):
                pass
        return None

    def _update_cells(
        self, ws, target_row: int, col_map: dict[str, int], updates: dict[str, Any]
    ) -> None:
        """Обновляет ячейки в строке."""
        for field, col_idx in col_map.items():
            value = updates.get(field)
            if value is not None:
                if field.startswith("margin"):
                    value = round(float(value), 2)
                ws.cell(target_row, col_idx, value=value)

    # ------------------------------------------------------------------
    # Приватные вспомогательные методы
    # ------------------------------------------------------------------

    @staticmethod
    def _parse_strategy_value(value) -> StrategyType:
        """
        Преобразует значение стратегии из Excel в StrategyType enum.

        Поддерживает:
            - числа 1, 2, 3,
            - текстовые варианты: 'ниже', 'выше', 'равная',
            - любые регистры.

        Args:
            value: Значение из ячейки (строка, число, None).

        Returns:
            StrategyType: Соответствующий тип стратегии (по умолчанию EQUAL).
        """
        return parse_strategy_value(value)

    def _parse_intervals_with_validation(
        self, row: pd.Series, columns: pd.Index
    ) -> tuple[list[StrategyInterval], list[str]]:
        """
        Парсит интервалы стратегий из строки Excel с валидацией.

        Args:
            row: Строка DataFrame.
            columns: Индекс колонок (для поиска по имени).

        Returns:
            Кортеж (список StrategyInterval, список предупреждений).
        """
        intervals = []
        warnings = []

        for i in range(1, settings.SCHEDULE_INTERVALS_COUNT + 1):
            time_col = self._find_column(columns, [f"интервал {i}", f"промежуток {i}"])
            if not time_col:
                continue

            time_val = row.get(time_col)
            if pd.isna(time_val) or not str(time_val).strip():
                continue

            time_range = str(time_val).strip()
            if "-" not in time_range:
                warnings.append(
                    f"Интервал {i}: неверный формат '{time_range}', ожидается ЧЧ:ММ-ЧЧ:ММ"
                )
                continue

            start, end = time_range.split("-", 1)
            start, end = start.strip(), end.strip()

            # Простая проверка формата HH:MM
            if not (
                len(start) == TIME_FORMAT_LENGTH
                and start[2] == ":"
                and start[:2].isdigit()
                and start[3:].isdigit()
            ):
                warnings.append(f"Интервал {i}: некорректное время начала '{start}'")
            if not (
                len(end) == TIME_FORMAT_LENGTH
                and end[2] == ":"
                and end[:2].isdigit()
                and end[3:].isdigit()
            ):
                warnings.append(f"Интервал {i}: некорректное время окончания '{end}'")

            strategy_col = self._find_column(columns, [f"стратегия {i}", f"стратеги {i}"])
            strategy_val = row.get(strategy_col) if strategy_col else None
            strategy = self._parse_strategy_value(strategy_val)

            percent_col = self._find_column(columns, [f"процент {i}", f"percent_{i}"])
            percent = 0.0
            if percent_col:
                percent_val = row.get(percent_col)
                if pd.notna(percent_val):
                    try:
                        percent = float(percent_val)
                        if strategy in (StrategyType.BELOW, StrategyType.ABOVE) and (
                            percent < 0 or percent > PERCENT_MAX
                        ):
                            warnings.append(
                                f"Интервал {i}: процент {percent} выходит за пределы 0-100, используется 0"
                            )
                            percent = 0.0
                    except Exception:
                        warnings.append(
                            f"Интервал {i}: процент '{percent_val}' не число, используется 0"
                        )

            intervals.append(
                StrategyInterval(start=start, end=end, strategy_type=strategy, percent=percent)
            )

        return intervals, warnings

    def _find_column(self, columns: pd.Index, candidates: list[str]) -> str | None:
        """Ищет колонку по одному из возможных имён."""
        for cand in candidates:
            if cand in columns:
                return cand
        return None

    def _get_float(
        self, row: pd.Series, columns: pd.Index, candidates: list[str], default: float
    ) -> float:
        """Извлекает числовое значение из ячейки по имени колонки."""
        col = self._find_column(columns, candidates)
        if col:
            val = row.get(col)
            if pd.notna(val):
                try:
                    return float(val)
                except Exception:
                    pass
        return default

    def build_excel_updates(
        self,
        product: ProductInfo,
        result: PriceCalculationResult,
        marginality_week: float,
        marginality_month: float,
        old_price_update: int | None,
    ) -> dict[str, Any]:
        """
        Формирует словарь обновлений для Excel на основе результатов расчёта.

        Args:
            product: Объект товара.
            result: Результат расчёта цены и маржинальности.
            marginality_week: Средняя маржинальность за неделю.
            marginality_month: Средняя маржинальность за месяц.
            old_price_update: Значение old_price для записи (или None).

        Returns:
            Словарь с полями: current_price, min_price, margin, margin_week,
            margin_month, old_price (если передан).
        """
        discount_coef = result.log_details.get("discount_coef", 1.0)
        real_price = result.result_target_price * discount_coef
        current_price_excel = int(round(real_price))
        min_price_excel = int(round(product.min_price))

        updates = {
            "current_price": current_price_excel,
            "min_price": min_price_excel,
            "margin": result.marginality,
            "margin_week": marginality_week,
            "margin_month": marginality_month,
        }
        if old_price_update is not None:
            updates["old_price"] = old_price_update
        return updates
