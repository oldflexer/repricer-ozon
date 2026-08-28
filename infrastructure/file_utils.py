import logging
import shutil
import time
from pathlib import Path
from typing import Any, Dict, Tuple

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

LOCK_WAIT_TIMEOUT = 60


def wait_for_excel_available(file_path: Path, timeout: int = 60) -> bool:
    """
    Wait until the Excel file becomes available (can be opened for writing).
    Returns True if file becomes available within timeout, False otherwise.
    """
    start = time.time()
    while time.time() - start < timeout:
        try:
            # Try to open the file in read mode to check if it's locked
            # Using 'r' mode to avoid truncating the file
            with file_path.open("r") as _:
                pass
            # If we can open it, file is available
            return True
        except PermissionError:
            # File is locked, wait and retry
            time.sleep(2)
        except Exception:
            # Other error (e.g., file doesn't exist), return False
            return False
    return False


def save_safely(
    updates: Dict[Tuple[int, int], Any], file_path: Path, max_retries: int = 3
) -> None:
    """
    Safely updates an Excel file with retry logic.
    Creates a backup, applies updates, saves to temp file, then replaces original.
    """
    last_error: Exception | None = None
    for attempt in range(max_retries):
        temp_path: Path | None = None
        try:
            # Create a backup
            backup_path = file_path.with_suffix(file_path.suffix + ".backup")
            shutil.copy2(file_path, backup_path)

            # Load the workbook
            wb = load_workbook(file_path)
            ws = wb.active
            if ws is None:
                raise ValueError("Нет активного листа")

            # Apply updates
            for (row, col), value in updates.items():
                ws.cell(row=row, column=col, value=value)

            # Save to a temporary file
            temp_path = file_path.with_suffix(file_path.suffix + ".tmp")
            wb.save(temp_path)
            wb.close()
            temp_path.replace(file_path)
            logger.info(f"Файл {file_path} успешно сохранён (попытка {attempt})")
            return
        except Exception as e:
            last_error = e
            logger.warning(f"Ошибка сохранения (попытка {attempt}/{max_retries}): {e}")
            if temp_path is not None and temp_path.exists():
                temp_path.unlink()
            if attempt < max_retries - 1:
                time.sleep(1)
    logger.error(f"Не удалось сохранить файл {file_path} после {max_retries} попыток: {last_error}")
    raise last_error or RuntimeError("Ошибка сохранения файла")
