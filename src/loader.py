import pandas as pd
from pathlib import Path
from typing import List, Dict, Optional, Any
import logging
import json
import re

logger = logging.getLogger(__name__)


class DataLoader:
    """Загрузка данных из Excel-файла с таблицей товаров (новая структура)"""

    COLUMN_MAPPING = {
        'sku': ['sku', 'артикул', 'article', 'id', 'offer_id'],
        'product_name': ['название', 'name', 'товар', 'product_name'],
        'cost_price': ['себестоимость', 'cost_price', 'cost'],
        'min_price': ['цена риц', 'минимальная цена', 'min_price', 'min'],
        'current_price': ['ваша цена', 'current_price', 'price'],
        'competitor_urls': ['конкурент', 'ссылки_конкурентов'],
    }

    COMPETITOR_COLUMNS_COUNT = 5
    SCHEDULE_INTERVALS_COUNT = 4

    def __init__(self, file_path: Path):
        self.file_path = file_path

    def load(self) -> List[Dict[str, Any]]:
        if not self.file_path.exists():
            logger.error(f"Файл {self.file_path} не найден")
            return []

        if self.file_path.suffix.lower() != '.xlsx':
            logger.error(f"Неподдерживаемый формат файла: {self.file_path.suffix}. Используйте .xlsx")
            return []

        df = pd.read_excel(self.file_path, engine='openpyxl', dtype=str)
        df.columns = df.columns.str.lower().str.strip()

        # Нормализация SKU
        for sku_col in ['sku', 'артикул', 'article', 'id', 'offer_id']:
            if sku_col in df.columns:
                df[sku_col] = df[sku_col].str.strip()
                break

        products = []
        for _, row in df.iterrows():
            product = self._parse_row(row, df.columns)
            if product:
                products.append(product)

        logger.info(f"Загружено {len(products)} товаров")
        return products

    def _parse_row(self, row: pd.Series, columns: pd.Index) -> Optional[Dict[str, Any]]:
        product = {}

        # 1. Простые поля
        for std_name, synonyms in self.COLUMN_MAPPING.items():
            if std_name == 'competitor_urls':
                continue
            value = None
            for syn in synonyms:
                if syn in columns:
                    val = row.get(syn)
                    if pd.notna(val):
                        value = val
                        break
            product[std_name] = value

        if not product.get('sku'):
            logger.warning("Пропущена строка без SKU")
            return None

        # 2. Ссылки конкурентов
        competitor_urls = []
        for i in range(1, self.COMPETITOR_COLUMNS_COUNT + 1):
            col_candidates = [f'конкурент {i}', f'конкурент{i}', f'конкурент_{i}']
            for col in col_candidates:
                if col in columns:
                    url = row.get(col)
                    if pd.notna(url) and str(url).strip():
                        competitor_urls.append(str(url).strip())
                    break
        product['competitor_urls'] = competitor_urls

        # 3. Интервалы стратегий
        intervals = []
        for i in range(1, self.SCHEDULE_INTERVALS_COUNT + 1):
            time_col = self._find_column(columns, [f'интервал {i}', f'промежуток {i}'])
            strategy_col = self._find_column(columns, [f'стратегия {i}', f'стратеги {i}'])
            percent_col = self._find_column(columns, [f'процент {i}', f'percent_{i}'])

            if not time_col:
                continue

            time_val = row.get(time_col)
            if pd.isna(time_val) or not str(time_val).strip():
                continue

            time_range = str(time_val).strip()
            if '-' not in time_range:
                logger.warning(f"Неверный формат интервала '{time_range}'")
                continue

            start, end = time_range.split('-', 1)
            start = start.strip()
            end = end.strip()

            strategy_val = row.get(strategy_col) if strategy_col else None
            percent_val = row.get(percent_col) if percent_col else None

            try:
                strategy = int(float(strategy_val)) if pd.notna(strategy_val) else 3
            except:
                strategy = 3

            try:
                percent = float(percent_val) if pd.notna(percent_val) else 0.0
            except:
                percent = 0.0

            intervals.append({
                'start': start,
                'end': end,
                'strategy': strategy,
                'percent': percent
            })

        if not intervals:
            base_strategy = self._get_int(row, columns, ['стратегия', 'strategy'], 3)
            base_percent = self._get_float(row, columns, ['процент', 'percent'], 0.0)
            intervals.append({
                'start': '00:00',
                'end': '23:59',
                'strategy': base_strategy,
                'percent': base_percent
            })

        product['intervals'] = intervals

        # 4. Числовые поля
        for field in ['cost_price', 'min_price', 'current_price']:
            val = product.get(field)
            if val is not None:
                try:
                    product[field] = float(val)
                except:
                    product[field] = 0.0
            else:
                product[field] = 0.0

        # product_id и offer_id будут заполнены позже из Ozon API
        product['product_id'] = None
        product['offer_id'] = None

        return product

    def _find_column(self, columns: pd.Index, candidates: List[str]) -> Optional[str]:
        for cand in candidates:
            if cand in columns:
                return cand
        return None

    def _get_int(self, row: pd.Series, columns: pd.Index, candidates: List[str], default: int) -> int:
        col = self._find_column(columns, candidates)
        if col:
            val = row.get(col)
            if pd.notna(val):
                try:
                    return int(float(val))
                except:
                    pass
        return default

    def _get_float(self, row: pd.Series, columns: pd.Index, candidates: List[str], default: float) -> float:
        col = self._find_column(columns, candidates)
        if col:
            val = row.get(col)
            if pd.notna(val):
                try:
                    return float(val)
                except:
                    pass
        return default

    def update_product_in_file(self, sku: str, updates: Dict[str, Any]) -> bool:
        """Обновляет поля товара в Excel-файле (цена и маржа)."""
        try:
            from openpyxl import load_workbook
            wb = load_workbook(self.file_path)
            ws = wb.active
            if ws is None:
                logger.error("Не удалось получить активный лист")
                return False

            header_row = 1
            col_map = {}
            target_columns = {
                'current_price': ['ваша цена', 'current_price', 'price'],
                'margin': ['маржинальность', 'маржа', 'margin'],
                'margin_week': ['маржинальность за неделю', 'margin_week'],
                'margin_month': ['маржинальность за месяц', 'margin_month'],
            }
            for field, synonyms in target_columns.items():
                for col_idx, cell in enumerate(ws[header_row], start=1):
                    if cell.value and str(cell.value).lower().strip() in synonyms:
                        col_map[field] = col_idx
                        break

            sku_col = None
            for col_idx, cell in enumerate(ws[header_row], start=1):
                if cell.value and str(cell.value).lower().strip() in ['sku', 'артикул', 'offer_id']:
                    sku_col = col_idx
                    break
            if sku_col is None:
                logger.error("Не найдена колонка 'SKU'")
                return False

            target_row = None
            for row_idx in range(2, ws.max_row + 1):
                sku_cell = ws.cell(row_idx, sku_col)
                cell_value = sku_cell.value
                if cell_value is None:
                    continue
                cell_str = str(cell_value).strip()
                # Ищем точное совпадение SKU (может быть строкой или числом)
                if cell_str == str(sku):
                    target_row = row_idx
                    break
                # Также пробуем сравнить как числа (для SKU без ведущих нулей)
                try:
                    if int(float(cell_str)) == int(float(sku)):
                        target_row = row_idx
                        break
                except (ValueError, TypeError):
                    pass

            if target_row is None:
                logger.warning(f"SKU {sku} не найден в файле")
                return False

            for field, col_idx in col_map.items():
                value = updates.get(field)
                if value is not None:
                    if field.startswith('margin'):
                        value = round(float(value), 2)
                    ws.cell(target_row, col_idx, value=value)

            wb.save(self.file_path)
            logger.info(f"Обновлены поля {list(updates.keys())} для SKU {sku}")
            return True
        except Exception as e:
            logger.error(f"Ошибка обновления Excel: {e}")
            return False